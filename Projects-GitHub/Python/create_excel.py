#!/usr/bin/env python3
"""
Excel Workbook Generator - Complete market analysis workbook
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import json

print("=" * 80)
print("GENERATING EXCEL WORKBOOK")
print("=" * 80)

# Load all data
df = pd.read_csv('sales_data_processed.csv')
df['Sale_Date'] = pd.to_datetime(df['Sale_Date'])
stats_df = pd.read_csv('statistics_summary.csv')
adjustment_df = pd.read_csv('adjustment_analysis.csv')
adjustment_df['Sale_Date'] = pd.to_datetime(adjustment_df['Sale_Date'])

with open('validation_info.json', 'r') as f:
    validation_info = json.load(f)

with open('adjustment_summary.json', 'r') as f:
    adjustment_summary = json.load(f)

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
subheader_font = Font(bold=True, size=10)
center_alignment = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# SHEET 1: SUMMARY REPORT
# ============================================================================

print("\n[Sheet 1] Creating Summary Report...")

ws1 = wb.create_sheet("Summary Report")

# Title
ws1['A1'] = "REAL PROPERTY MARKET ANALYSIS"
ws1['A1'].font = Font(bold=True, size=14)
ws1['A2'] = "Version 2.3.1 - Comprehensive Statistical Analysis"
ws1['A2'].font = Font(italic=True, size=10)

# Subject Information
row = 4
ws1[f'A{row}'] = "SUBJECT PROPERTY INFORMATION"
ws1[f'A{row}'].font = subheader_font
ws1[f'A{row}'].fill = subheader_fill
row += 1

ws1[f'A{row}'] = "Address:"
ws1[f'B{row}'] = "528 S. Taper Ave, Compton, CA 90220"
row += 1
ws1[f'A{row}'] = "Living Area:"
ws1[f'B{row}'] = "951 SF"
row += 1
ws1[f'A{row}'] = "Bedrooms/Bathrooms:"
ws1[f'B{row}'] = "2 / 1"
row += 1
ws1[f'A{row}'] = "Appraiser:"
ws1[f'B{row}'] = "Craig Gilbert, ASA, SRA, CRP, Certified General Appraiser"
row += 1
ws1[f'A{row}'] = "File Number:"
ws1[f'B{row}'] = "25-060"
row += 1
ws1[f'A{row}'] = "Date of Value:"
ws1[f'B{row}'] = "December 31, 2013"
row += 2

# Data Coverage
ws1[f'A{row}'] = "DATA COVERAGE ANALYSIS"
ws1[f'A{row}'].font = subheader_font
ws1[f'A{row}'].fill = subheader_fill
row += 1

ws1[f'A{row}'] = "Actual Data Coverage:"
ws1[f'B{row}'] = f"{validation_info['actual_coverage_months']:.1f} months"
row += 1
ws1[f'A{row}'] = "Sales Date Range:"
ws1[f'B{row}'] = f"{pd.to_datetime(validation_info['earliest_sale']).strftime('%m/%d/%Y')} to {pd.to_datetime(validation_info['latest_sale']).strftime('%m/%d/%Y')}"
row += 1
ws1[f'A{row}'] = "Total Sales Analyzed:"
ws1[f'B{row}'] = len(df)
row += 1
ws1[f'A{row}'] = "Valid Time Periods:"
ws1[f'B{row}'] = len(validation_info['valid_periods'])
row += 1
ws1[f'A{row}'] = "Omitted Time Periods:"
ws1[f'B{row}'] = len(validation_info['omitted_periods'])
row += 2

# Market Trend
ws1[f'A{row}'] = "MARKET TREND ANALYSIS"
ws1[f'A{row}'].font = subheader_font
ws1[f'A{row}'].fill = subheader_fill
row += 1

trend_results = validation_info['trend_results']
ws1[f'A{row}'] = "Market Trend:"
ws1[f'B{row}'] = trend_results['market_trend']
ws1[f'B{row}'].font = Font(bold=True)
row += 1
ws1[f'A{row}'] = "Monthly Change:"
ws1[f'B{row}'] = f"{trend_results['monthly_price_change_pct']:+.2f}% per month"
row += 1
ws1[f'A{row}'] = "Daily Change:"
ws1[f'B{row}'] = f"${trend_results['daily_price_change']:+.2f} per day"
row += 2

# Adjustment Analysis Summary
ws1[f'A{row}'] = "COMPARABLE SALES ADJUSTMENT ANALYSIS"
ws1[f'A{row}'].font = subheader_font
ws1[f'A{row}'].fill = subheader_fill
row += 1

ws1[f'A{row}'] = "Unadjusted Median Price:"
ws1[f'B{row}'] = f"${adjustment_summary['unadjusted_median']:,.0f}"
row += 1
ws1[f'A{row}'] = "Adjusted Median Price:"
ws1[f'B{row}'] = f"${adjustment_summary['adjusted_median']:,.0f}"
ws1[f'B{row}'].font = Font(bold=True)
row += 1
ws1[f'A{row}'] = "Net Adjustment Impact:"
ws1[f'B{row}'] = f"${adjustment_summary['adjustment_impact']:+,.0f} ({adjustment_summary['adjustment_impact_pct']:+.2f}%)"
row += 1
ws1[f'A{row}'] = "Time Adjustments Applied:"
ws1[f'B{row}'] = f"{adjustment_summary['time_adjustments_applied']} of {adjustment_summary['n_comparables']}"
row += 1
ws1[f'A{row}'] = "SF Adjustments Applied:"
ws1[f'B{row}'] = f"{adjustment_summary['sf_adjustments_applied']} of {adjustment_summary['n_comparables']}"
row += 2

ws1[f'A{row}'] = "INDICATED VALUE:"
ws1[f'A{row}'].font = Font(bold=True, size=12)
ws1[f'B{row}'] = "[To be determined by appraiser]"
ws1[f'B{row}'].font = Font(italic=True, size=11)
ws1[f'B{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Adjust column widths
ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 50

print("  ✓ Summary Report created")

# ============================================================================
# SHEET 2: STATISTICS BY TIME PERIOD
# ============================================================================

print("[Sheet 2] Creating Statistics by Time Period...")

ws2 = wb.create_sheet("Statistics by Period")

# Title
ws2['A1'] = "STATISTICS BY TIME PERIOD"
ws2['A1'].font = Font(bold=True, size=12)
ws2.merge_cells('A1:J1')
ws2['A1'].alignment = center_alignment

# Headers
headers = ['Period', 'Months', 'N Sales', 'Absorption Rate', 'Price Mean', 'Price Median', 
           'Price Std Dev', '$/SF Mean', '$/SF Median', '$/SF Std Dev']
for col_num, header in enumerate(headers, 1):
    cell = ws2.cell(row=3, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = border

# Data
for row_num, row_data in enumerate(dataframe_to_rows(stats_df, index=False, header=False), 4):
    for col_num, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = border
        
        # Format numbers
        if col_num == 4:  # Absorption Rate
            cell.number_format = '0.00'
        elif col_num in [5, 6, 7]:  # Prices
            cell.number_format = '$#,##0'
        elif col_num in [8, 9, 10]:  # Price per SF
            cell.number_format = '$#,##0.00'

# Adjust column widths
for col in range(1, 11):
    ws2.column_dimensions[chr(64 + col)].width = 14

print("  ✓ Statistics by Period created")

# ============================================================================
# SHEET 3: DATA COVERAGE ANALYSIS
# ============================================================================

print("[Sheet 3] Creating Data Coverage Analysis...")

ws3 = wb.create_sheet("Data Coverage Analysis")

ws3['A1'] = "DATA COVERAGE VALIDATION"
ws3['A1'].font = Font(bold=True, size=12)

row = 3
ws3[f'A{row}'] = "Coverage Period:"
ws3[f'B{row}'] = f"{validation_info['actual_coverage_months']:.1f} months"
row += 1
ws3[f'A{row}'] = "Earliest Sale:"
ws3[f'B{row}'] = pd.to_datetime(validation_info['earliest_sale']).strftime('%m/%d/%Y')
row += 1
ws3[f'A{row}'] = "Latest Sale:"
ws3[f'B{row}'] = pd.to_datetime(validation_info['latest_sale']).strftime('%m/%d/%Y')
row += 1
ws3[f'A{row}'] = "Date of Value:"
ws3[f'B{row}'] = "12/31/2013"
row += 3

# Valid Periods
ws3[f'A{row}'] = "VALID TIME PERIODS (INCLUDED IN ANALYSIS)"
ws3[f'A{row}'].font = subheader_font
ws3[f'A{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
row += 1

ws3[f'A{row}'] = "Period"
ws3[f'B{row}'] = "Months"
ws3[f'C{row}'] = "Sales Count"
ws3[f'D{row}'] = "Status"
for col in ['A', 'B', 'C', 'D']:
    ws3[f'{col}{row}'].font = header_font
    ws3[f'{col}{row}'].fill = header_fill
row += 1

for period in validation_info['valid_periods']:
    ws3[f'A{row}'] = period['name']
    ws3[f'B{row}'] = period['months']
    ws3[f'C{row}'] = period['sales_count']
    ws3[f'D{row}'] = "✓ Included"
    row += 1

row += 1

# Omitted Periods
ws3[f'A{row}'] = "OMITTED TIME PERIODS (EXCLUDED FROM ANALYSIS)"
ws3[f'A{row}'].font = subheader_font
ws3[f'A{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
row += 1

ws3[f'A{row}'] = "Period"
ws3[f'B{row}'] = "Months"
ws3[f'C{row}'] = "Reason"
for col in ['A', 'B', 'C']:
    ws3[f'{col}{row}'].font = header_font
    ws3[f'{col}{row}'].fill = header_fill
row += 1

for period in validation_info['omitted_periods']:
    ws3[f'A{row}'] = period['name']
    ws3[f'B{row}'] = period['months']
    ws3[f'C{row}'] = period['reason']
    row += 1

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 50

print("  ✓ Data Coverage Analysis created")

# ============================================================================
# SHEET 4: COMPARABLE SALES ADJUSTMENTS
# ============================================================================

print("[Sheet 4] Creating Comparable Sales Adjustments...")

ws4 = wb.create_sheet("Comp Sales Adjustments")

ws4['A1'] = "COMPARABLE SALES ADJUSTMENT ANALYSIS"
ws4['A1'].font = Font(bold=True, size=12)
ws4.merge_cells('A1:M1')
ws4['A1'].alignment = center_alignment

# Info
ws4['A2'] = f"Subject Living Area: {adjustment_summary['subject_living_area']:,} SF"
ws4['A3'] = f"Time Threshold: {adjustment_summary['time_threshold_days']} days | SF Threshold: {adjustment_summary['sf_threshold_pct']}%"

# Headers
headers = ['Sale Date', 'Address', 'Sale Price', 'Living Area', 'Days Diff', 
           'Time Adj $', 'Time Adj %', 'SF Diff', 'SF Adj $', 
           'Total Adj $', 'Total Adj %', 'Adjusted Price', '$/SF Adjusted']
for col_num, header in enumerate(headers, 1):
    cell = ws4.cell(row=5, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = border

# Data
for row_num, row_data in enumerate(dataframe_to_rows(adjustment_df, index=False, header=False), 6):
    for col_num, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_num, column=col_num)
        
        # Format sale date
        if col_num == 1:
            cell.value = pd.to_datetime(value).strftime('%m/%d/%Y')
        # Format address
        elif col_num == 2:
            addr_parts = [adjustment_df.iloc[row_num-6]['Street_Number'], 
                          adjustment_df.iloc[row_num-6]['Street_Name']]
            cell.value = f"{int(addr_parts[0])} {addr_parts[1]}"
        else:
            cell.value = value
        
        cell.border = border
        
        # Format numbers
        if col_num in [3, 6, 9, 10, 12]:  # Dollar amounts
            cell.number_format = '$#,##0'
        elif col_num in [7, 11]:  # Percentages
            cell.number_format = '0.00%'
        elif col_num == 13:  # Price per SF
            cell.number_format = '$#,##0.00'

# Summary statistics below table
summary_row = len(adjustment_df) + 8
ws4[f'A{summary_row}'] = "SUMMARY STATISTICS"
ws4[f'A{summary_row}'].font = subheader_font
ws4[f'A{summary_row}'].fill = subheader_fill
summary_row += 1

ws4[f'A{summary_row}'] = "Unadjusted Median:"
ws4[f'B{summary_row}'] = adjustment_summary['unadjusted_median']
ws4[f'B{summary_row}'].number_format = '$#,##0'
summary_row += 1

ws4[f'A{summary_row}'] = "Adjusted Median:"
ws4[f'B{summary_row}'] = adjustment_summary['adjusted_median']
ws4[f'B{summary_row}'].number_format = '$#,##0'
ws4[f'B{summary_row}'].font = Font(bold=True)
ws4[f'B{summary_row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
summary_row += 1

ws4[f'A{summary_row}'] = "Net Adjustment:"
ws4[f'B{summary_row}'] = adjustment_summary['adjustment_impact']
ws4[f'B{summary_row}'].number_format = '$#,##0'
summary_row += 2

ws4[f'A{summary_row}'] = "INDICATED VALUE:"
ws4[f'A{summary_row}'].font = Font(bold=True, size=11)
ws4[f'B{summary_row}'] = "[Appraiser to determine]"
ws4[f'B{summary_row}'].font = Font(italic=True)
ws4[f'B{summary_row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Adjust column widths
for col in range(1, 14):
    ws4.column_dimensions[chr(64 + col)].width = 12

print("  ✓ Comparable Sales Adjustments created")

# ============================================================================
# SHEET 5: ALL SALES DATA
# ============================================================================

print("[Sheet 5] Creating All Sales Data...")

ws5 = wb.create_sheet("All Sales Data")

ws5['A1'] = "COMPLETE SALES DATA"
ws5['A1'].font = Font(bold=True, size=12)

# Select relevant columns
sales_cols = ['Sale_Date', 'Street_Number', 'Street_Name', 'Sale_Price', 
              'Living_Area', 'Price_Per_SF', 'Bedrooms', 'Bathrooms', 'DOM', 'CDOM']
sales_export = df[sales_cols].copy()
sales_export['Sale_Date'] = sales_export['Sale_Date'].dt.strftime('%m/%d/%Y')

# Headers
headers = ['Sale Date', 'Street #', 'Street Name', 'Sale Price', 'Living Area', 
           '$/SF', 'Beds', 'Baths', 'DOM', 'CDOM']
for col_num, header in enumerate(headers, 1):
    cell = ws5.cell(row=3, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

# Data
for row_num, row_data in enumerate(dataframe_to_rows(sales_export, index=False, header=False), 4):
    for col_num, value in enumerate(row_data, 1):
        cell = ws5.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = border
        
        if col_num == 4:  # Sale Price
            cell.number_format = '$#,##0'
        elif col_num == 6:  # Price per SF
            cell.number_format = '$#,##0.00'

for col in range(1, 11):
    ws5.column_dimensions[chr(64 + col)].width = 13

print("  ✓ All Sales Data created")

# ============================================================================
# SAVE WORKBOOK
# ============================================================================

output_file = 'Market_Analysis_528_S_Taper_Ave.xlsx'
wb.save(output_file)

print("\n" + "=" * 80)
print("✓ EXCEL WORKBOOK CREATED SUCCESSFULLY")
print("=" * 80)
print(f"File: {output_file}")
print(f"Sheets: {len(wb.sheetnames)}")
for i, sheet in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {sheet}")
print("=" * 80)
